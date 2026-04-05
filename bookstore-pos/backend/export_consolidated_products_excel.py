from __future__ import annotations

import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "bookstore-pos" / "backend" / "bookstore.db"
UPDATED_PRICES_PATH = ROOT / "productos_actualizados_con_precios.xlsx"
EXPORTED_PRODUCTS_PATH = ROOT / "productos_exportados.xlsx"
OUTPUT_PATH = ROOT / "base_productos_precios_actualizada.xlsx"


def normalize_name(value: Any) -> str:
    text = "" if value is None else str(value)
    normalized = unicodedata.normalize("NFKD", text)
    without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    cleaned = []
    previous_was_space = False
    for ch in without_accents.lower().strip():
        if ch.isalnum():
            cleaned.append(ch)
            previous_was_space = False
            continue
        if not previous_was_space:
            cleaned.append(" ")
            previous_was_space = True
    return "".join(cleaned).strip()


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        width = max((len(str(value)) for value in values), default=10)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 60)


def load_db_products() -> list[dict[str, Any]]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT
                id,
                sku,
                name,
                category,
                price,
                cost,
                stock,
                stock_min,
                tax_rate,
                tax_included,
                tags,
                sale_price,
                author,
                publisher,
                isbn,
                barcode,
                shelf_location,
                desired_margin,
                unit_cost
            FROM products
            ORDER BY name COLLATE NOCASE
            """
        ).fetchall()
    finally:
        conn.close()
    return [dict(row) for row in rows]


def load_updated_price_rows() -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    workbook = load_workbook(UPDATED_PRICES_PATH, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, Any]] = []
    indexed: dict[str, list[dict[str, Any]]] = {}
    try:
        for source_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None for value in row):
                continue
            record = {
                "source_row": source_row_number,
                "description": row[1],
                "quantity": row[2],
                "price": row[3],
                "cost": row[4],
            }
            rows.append(record)
            key = normalize_name(record["description"])
            indexed.setdefault(key, []).append(record)
    finally:
        workbook.close()
    return rows, indexed


def load_exported_products() -> list[dict[str, Any]]:
    workbook = load_workbook(EXPORTED_PRODUCTS_PATH, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, Any]] = []
    try:
        headers = [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            rows.append(dict(zip(headers, row)))
    finally:
        workbook.close()
    return rows


def build_consolidated_rows(
    db_products: list[dict[str, Any]],
    updated_by_name: dict[str, list[dict[str, Any]]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    consolidated: list[dict[str, Any]] = []
    matched_keys: set[str] = set()
    duplicate_keys: list[str] = []

    for product in db_products:
        key = normalize_name(product["name"])
        updated_candidates = updated_by_name.get(key, [])
        row = dict(product)
        row["source"] = "sqlite"
        row["match_status"] = "sin_cambios"
        row["updated_source_row"] = None

        if updated_candidates:
            matched_keys.add(key)
            if len(updated_candidates) > 1:
                duplicate_keys.append(product["name"])
            update = updated_candidates[-1]
            row["price"] = update["price"] if update["price"] is not None else row["price"]
            row["cost"] = update["cost"] if update["cost"] is not None else row["cost"]
            row["stock"] = int(update["quantity"]) if update["quantity"] is not None else row["stock"]
            row["sale_price"] = row["price"]
            row["unit_cost"] = row["cost"]
            row["source"] = "sqlite+excel_actualizado"
            row["match_status"] = "actualizado_por_nombre"
            row["updated_source_row"] = update["source_row"]

        consolidated.append(row)

    unmatched_updates = []
    for key, records in updated_by_name.items():
        if key in matched_keys:
            continue
        unmatched_updates.extend(records)

    unmatched_updates.sort(key=lambda item: normalize_name(item["description"]))
    duplicate_keys.sort()
    return consolidated, unmatched_updates, duplicate_keys


def build_export_only_rows(db_products: list[dict[str, Any]], exported_products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    db_skus = {str(product["sku"]).strip() for product in db_products}
    export_only = []
    for row in exported_products:
        sku = str(row.get("SKU") or "").strip()
        if not sku or sku in db_skus:
            continue
        export_only.append(row)
    export_only.sort(key=lambda item: str(item.get("NAME") or "").lower())
    return export_only


def write_sheet(worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    autosize_columns(worksheet)


def main() -> None:
    db_products = load_db_products()
    updated_rows, updated_by_name = load_updated_price_rows()
    exported_products = load_exported_products()
    consolidated_rows, unmatched_updates, duplicate_match_names = build_consolidated_rows(db_products, updated_by_name)
    export_only_rows = build_export_only_rows(db_products, exported_products)

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    summary_sheet.append(["campo", "valor"])
    summary_sheet.append(["generado_en", datetime.now().isoformat(timespec="seconds")])
    summary_sheet.append(["archivo_salida", str(OUTPUT_PATH.name)])
    summary_sheet.append(["productos_en_db", len(db_products)])
    summary_sheet.append(["filas_excel_actualizado", len(updated_rows)])
    summary_sheet.append(
        [
            "productos_actualizados_por_nombre",
            sum(1 for row in consolidated_rows if row["match_status"] == "actualizado_por_nombre"),
        ]
    )
    summary_sheet.append(["filas_excel_pendientes_revision", len(unmatched_updates)])
    summary_sheet.append(["filas_exportadas_no_vigentes", len(export_only_rows)])
    summary_sheet.append(["coincidencias_duplicadas_en_excel", len(duplicate_match_names)])
    autosize_columns(summary_sheet)

    import_headers = [
        "sku",
        "name",
        "category",
        "tags",
        "price",
        "cost",
        "stock",
        "stock_min",
        "author",
        "publisher",
        "isbn",
        "barcode",
        "shelf_location",
        "desired_margin",
        "unit_cost",
        "tax_rate",
        "tax_included",
        "sale_price",
    ]
    import_rows = [{header: row.get(header) for header in import_headers} for row in consolidated_rows]
    write_sheet(workbook.create_sheet("Importacion_Sistema"), import_headers, import_rows)

    consolidated_headers = [
        "id",
        "sku",
        "name",
        "category",
        "price",
        "cost",
        "stock",
        "stock_min",
        "tax_rate",
        "tax_included",
        "tags",
        "sale_price",
        "author",
        "publisher",
        "isbn",
        "barcode",
        "shelf_location",
        "desired_margin",
        "unit_cost",
        "source",
        "match_status",
        "updated_source_row",
    ]
    write_sheet(workbook.create_sheet("Catalogo_Consolidado"), consolidated_headers, consolidated_rows)

    pending_headers = ["source_row", "description", "quantity", "price", "cost"]
    write_sheet(workbook.create_sheet("Excel_Pendientes"), pending_headers, unmatched_updates)

    export_only_headers = [
        "ID",
        "SKU",
        "NAME",
        "CATEGORY",
        "PRICE",
        "COST",
        "STOCK",
        "STOCK_MIN",
        "TAX_RATE",
        "TAX_INCLUDED",
        "TAGS",
        "SALE_PRICE",
        "AUTHOR",
        "PUBLISHER",
        "ISBN",
        "BARCODE",
        "SHELF_LOCATION",
        "DESIRED_MARGIN",
        "UNIT_COST",
    ]
    write_sheet(workbook.create_sheet("Exportado_No_Vigente"), export_only_headers, export_only_rows)

    duplicate_sheet = workbook.create_sheet("Coincidencias_Duplicadas")
    duplicate_sheet.append(["name"])
    for name in duplicate_match_names:
        duplicate_sheet.append([name])
    autosize_columns(duplicate_sheet)

    workbook.save(OUTPUT_PATH)
    print(f"Archivo generado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
