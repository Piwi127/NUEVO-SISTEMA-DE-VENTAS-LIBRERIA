import csv
import json
import logging
import os
from datetime import datetime, timezone
from types import SimpleNamespace
from typing import Iterator
from time import perf_counter

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_event
from app.core.config import settings
from app.core.metrics import inventory_import_job_duration_seconds, inventory_import_jobs_total, inventory_import_rows_total
from app.core.stock import require_default_warehouse_id
import app.db.session as db_session
from app.models.inventory import InventoryImportJob, InventoryImportJobError
from app.services.inventory.stock_service import StockService

logger = logging.getLogger("bookstore.import")

REQUIRED_COLUMNS = {"sku", "name", "category", "price", "cost", "stock", "stock_min"}
ALLOWED_JOB_STATUSES = {"pending", "running", "success", "failed", "partial"}


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _safe_status(value: str) -> str:
    return value if value in ALLOWED_JOB_STATUSES else "failed"


def clamp_batch_size(batch_size: int | None) -> int:
    default_value = settings.inventory_import_default_batch_size
    max_value = settings.inventory_import_max_batch_size
    if batch_size is None:
        return default_value
    return min(max(1, int(batch_size)), max_value)


def _validate_headers(headers: list[str]) -> None:
    header_set = {header for header in headers if header}
    if not REQUIRED_COLUMNS.issubset(header_set):
        missing = ", ".join(sorted(REQUIRED_COLUMNS - header_set))
        raise ValueError(f"Faltan columnas: {missing}")


def _iter_csv_rows(file_path: str) -> Iterator[tuple[int, dict[str, str]]]:
    with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as handler:
        reader = csv.DictReader(handler)
        if not reader.fieldnames:
            raise ValueError("CSV sin encabezados")
        headers = [str(value).strip() for value in reader.fieldnames if value]
        _validate_headers(headers)
        for row_number, row in enumerate(reader, start=2):
            yield row_number, row


def _iter_xlsx_rows(file_path: str) -> Iterator[tuple[int, dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError("Instale openpyxl para XLSX") from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        workbook.close()
        raise ValueError("XLSX sin hoja activa")

    header_rows = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_rows:
        workbook.close()
        raise ValueError("XLSX sin encabezados")
    headers = [str(cell).strip() if cell is not None else "" for cell in header_rows[0]]
    if not any(headers):
        workbook.close()
        raise ValueError("XLSX sin encabezados")
    _validate_headers(headers)

    try:
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            payload = {headers[index]: row[index] for index in range(len(headers))}
            yield row_number, payload
    finally:
        workbook.close()


def iter_file_rows(file_path: str, file_type: str) -> Iterator[tuple[int, dict[str, object]]]:
    if file_type == "csv":
        yield from _iter_csv_rows(file_path)
        return
    if file_type == "xlsx":
        yield from _iter_xlsx_rows(file_path)
        return
    raise ValueError("Formato no soportado. Use CSV o XLSX")


class InventoryImportJobService:
    def __init__(self, db: AsyncSession, current_user, request_id: str | None = None):
        self.db = db
        self.current_user = current_user
        self.request_id = request_id

    async def create_job(self, *, filename: str, file_type: str, batch_size: int) -> InventoryImportJob:
        normalized_batch = clamp_batch_size(batch_size)
        job = InventoryImportJob(
            created_by=self.current_user.id,
            status="pending",
            filename=filename,
            file_type=file_type,
            request_id=self.request_id,
            batch_size=normalized_batch,
            total_rows=0,
            processed_rows=0,
            success_rows=0,
            error_rows=0,
            error_message=None,
            created_at=_now(),
            started_at=None,
            finished_at=None,
            updated_at=_now(),
        )
        self.db.add(job)
        await self.db.commit()
        await self.db.refresh(job)
        return job

    async def get_job(self, job_id: int) -> InventoryImportJob:
        result = await self.db.execute(select(InventoryImportJob).where(InventoryImportJob.id == job_id))
        job = result.scalar_one_or_none()
        if not job:
            raise HTTPException(status_code=404, detail="Job no encontrado")
        return job

    async def get_job_errors(self, job_id: int, *, limit: int = 500) -> tuple[InventoryImportJob, list[InventoryImportJobError]]:
        job = await self.get_job(job_id)
        safe_limit = min(max(limit, 1), 5000)
        result = await self.db.execute(
            select(InventoryImportJobError)
            .where(InventoryImportJobError.job_id == job_id)
            .order_by(InventoryImportJobError.id.asc())
            .limit(safe_limit)
        )
        return job, list(result.scalars().all())

    async def export_job_errors_csv(self, job_id: int) -> str:
        _, rows = await self.get_job_errors(job_id=job_id, limit=5000)
        from io import StringIO

        buffer = StringIO()
        csv_writer = csv.writer(buffer)
        csv_writer.writerow(["id", "row_number", "sku", "detail", "raw_data", "created_at"])
        for row in rows:
            csv_writer.writerow([row.id, row.row_number, row.sku or "", row.detail, row.raw_data or "", row.created_at.isoformat()])
        return buffer.getvalue()


async def _mark_job_running(session: AsyncSession, job: InventoryImportJob) -> None:
    now = _now()
    job.status = "running"
    job.started_at = now
    job.updated_at = now
    job.error_message = None
    await session.commit()


async def _record_row_error(
    session: AsyncSession,
    job: InventoryImportJob,
    *,
    row_number: int,
    row_data: dict[str, object],
    error_detail: str,
) -> None:
    job.error_rows += 1
    job.processed_rows += 1
    job.total_rows += 1
    job.updated_at = _now()
    session.add(
        InventoryImportJobError(
            job_id=job.id,
            row_number=row_number,
            sku=str(row_data.get("sku") or "").strip() or None,
            detail=error_detail,
            raw_data=json.dumps(row_data, ensure_ascii=False, default=str),
            created_at=_now(),
        )
    )
    inventory_import_rows_total.labels("error").inc()
    await session.commit()


async def _record_row_success(session: AsyncSession, job: InventoryImportJob) -> None:
    job.success_rows += 1
    job.processed_rows += 1
    job.total_rows += 1
    job.updated_at = _now()
    inventory_import_rows_total.labels("success").inc()
    await session.commit()


async def run_inventory_import_job(job_id: int, file_path: str) -> None:
    started_perf = perf_counter()
    final_status = "failed"
    request_id = ""
    try:
        async with db_session.AsyncSessionLocal() as session:
            job_result = await session.execute(select(InventoryImportJob).where(InventoryImportJob.id == job_id))
            job = job_result.scalar_one_or_none()
            if not job:
                return
            request_id = job.request_id or ""
            await _mark_job_running(session, job)
            service = StockService(session, SimpleNamespace(id=job.created_by))
            default_warehouse_id = await require_default_warehouse_id(session)
            try:
                for row_number, row in iter_file_rows(file_path, job.file_type):
                    try:
                        parsed = service.parse_import_row(row_number, row)
                    except ValueError as exc:
                        await _record_row_error(session, job, row_number=row_number, row_data=row, error_detail=str(exc))
                        continue
                    if parsed is None:
                        continue

                    try:
                        await service.upsert_import_row(
                            parsed,
                            default_warehouse_id=default_warehouse_id,
                            ref=f"IMPORT_JOB:{job.id}",
                        )
                        await _record_row_success(session, job)
                    except Exception as exc:
                        await session.rollback()
                        await _record_row_error(session, job, row_number=row_number, row_data=row, error_detail=str(exc))

                job.finished_at = _now()
                job.updated_at = _now()
                if job.processed_rows == 0 and job.error_rows > 0:
                    job.status = "failed"
                elif job.error_rows > 0:
                    job.status = "partial"
                else:
                    job.status = "success"
                final_status = _safe_status(job.status)
                await log_event(
                    session,
                    job.created_by,
                    "inventory_import_job",
                    "inventory_import_job",
                    str(job.id),
                    f"status={job.status};success={job.success_rows};errors={job.error_rows}",
                )
                await session.commit()
            except Exception as exc:
                await session.rollback()
                job.status = "failed"
                job.error_message = str(exc)
                job.finished_at = _now()
                job.updated_at = _now()
                final_status = "failed"
                await session.commit()
                logger.exception(
                    "Inventory import job failed request_id=%s job_id=%s error=%s",
                    request_id or "-",
                    job.id,
                    str(exc),
                )
            else:
                logger.info(
                    "Inventory import job completed request_id=%s job_id=%s status=%s processed=%s success=%s errors=%s",
                    request_id or "-",
                    job.id,
                    job.status,
                    job.processed_rows,
                    job.success_rows,
                    job.error_rows,
                )
    finally:
        duration = perf_counter() - started_perf
        inventory_import_job_duration_seconds.observe(duration)
        inventory_import_jobs_total.labels(final_status).inc()
        try:
            os.remove(file_path)
        except FileNotFoundError:
            pass
