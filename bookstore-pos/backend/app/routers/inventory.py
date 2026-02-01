import csv
from io import StringIO, BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import PlainTextResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_db, require_role
from app.models.inventory import StockMovement
from app.models.product import Product
from app.schemas.inventory import InventoryMovementCreate, StockMovementOut

router = APIRouter(prefix="/inventory", tags=["inventory"], dependencies=[Depends(require_role("admin", "stock"))])


REQUIRED_COLUMNS = {"sku", "name", "category", "price", "cost", "stock", "stock_min"}


@router.post("/movement", response_model=StockMovementOut, status_code=201)
async def create_movement(data: InventoryMovementCreate, db: AsyncSession = Depends(get_db)):
    if data.qty <= 0:
        raise HTTPException(status_code=400, detail="Cantidad invalida")
    if data.type not in {"IN", "ADJ"}:
        raise HTTPException(status_code=400, detail="Tipo invalido")
    result = await db.execute(select(Product).where(Product.id == data.product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    product.stock += data.qty
    movement = StockMovement(
        product_id=data.product_id,
        type=data.type,
        qty=data.qty,
        ref=data.ref,
    )
    db.add(movement)
    await db.commit()
    await db.refresh(movement)
    return movement


@router.get("/template")
async def download_template():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["sku", "name", "category", "price", "cost", "stock", "stock_min"])
    writer.writerow(["BK-001", "Libro ejemplo", "Ficcion", "25.90", "12.50", "10", "2"])
    return PlainTextResponse(output.getvalue(), media_type="text/csv")


@router.get("/template/xlsx")
async def download_template_xlsx():
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=400, detail="Instale openpyxl para XLSX")
    wb = Workbook()
    ws = wb.active
    ws.append(["sku", "name", "category", "price", "cost", "stock", "stock_min"])
    ws.append(["BK-001", "Libro ejemplo", "Ficcion", "25.90", "12.50", "10", "2"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_template.xlsx"},
    )


@router.post("/upload")
async def upload_inventory(file: UploadFile = File(...), db: AsyncSession = Depends(get_db), _=Depends(require_role("admin"))):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Archivo invalido")
    filename = file.filename.lower()
    rows: list[dict] = []

    if filename.endswith(".csv"):
        content = await file.read()
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="CSV sin encabezados")
        header = {h.strip() for h in reader.fieldnames if h}
        if not REQUIRED_COLUMNS.issubset(header):
            missing = ", ".join(sorted(REQUIRED_COLUMNS - header))
            raise HTTPException(status_code=400, detail=f"Faltan columnas: {missing}")
        rows = [r for r in reader]
    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(status_code=400, detail="Instale openpyxl para XLSX")
        data = await file.read()
        wb = load_workbook(filename=BytesIO(data))
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header = {h for h in headers if h}
        if not REQUIRED_COLUMNS.issubset(header):
            missing = ", ".join(sorted(REQUIRED_COLUMNS - header))
            raise HTTPException(status_code=400, detail=f"Faltan columnas: {missing}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: row[i] for i in range(len(headers))})
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o XLSX")

    if not rows:
        return {"ok": True, "count": 0}

    async with db.begin():
        for r in rows:
            sku = str(r.get("sku") or "").strip()
            name = str(r.get("name") or "").strip()
            if not sku or not name:
                continue
            category = str(r.get("category") or "").strip()
            price = float(r.get("price") or 0)
            cost = float(r.get("cost") or 0)
            stock = int(float(r.get("stock") or 0))
            stock_min = int(float(r.get("stock_min") or 0))

            result = await db.execute(select(Product).where(Product.sku == sku))
            product = result.scalar_one_or_none()
            if product:
                diff = stock - product.stock
                product.name = name
                product.category = category
                product.price = price
                product.cost = cost
                product.stock = stock
                product.stock_min = stock_min
                if diff != 0:
                    movement = StockMovement(
                        product_id=product.id,
                        type="ADJ",
                        qty=abs(diff),
                        ref="BULK_IMPORT",
                    )
                    db.add(movement)
            else:
                product = Product(
                    sku=sku,
                    name=name,
                    category=category,
                    price=price,
                    cost=cost,
                    stock=stock,
                    stock_min=stock_min,
                )
                db.add(product)
                await db.flush()
                if stock != 0:
                    movement = StockMovement(
                        product_id=product.id,
                        type="IN",
                        qty=stock,
                        ref="BULK_IMPORT",
                    )
                    db.add(movement)

    return {"ok": True, "count": len(rows)}


@router.get("/kardex/{product_id}", response_model=list[StockMovementOut])
async def get_kardex(product_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockMovement).where(StockMovement.product_id == product_id).order_by(StockMovement.id.desc()).limit(200)
    )
    return result.scalars().all()
