"""
Router de inventario.
Endpoints: POST /inventory/movement, /upload, /import-jobs, GET /inventory/kardex/{id}
"""

import base64
import csv
import os
import tempfile
from datetime import datetime
from io import BytesIO, StringIO

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.responses import PlainTextResponse, StreamingResponse
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.deps import get_current_user, get_db, require_permission, require_role
from app.models.inventory import InventoryImportJob, StockMovement
from app.schemas.inventory import (
    InventoryImportJobErrorOut,
    InventoryImportJobErrorListOut,
    InventoryImportJobOut,
    InventoryMovementCreate,
    KardexPageOut,
    StockMovementOut,
)
from app.services.inventory.import_jobs_service import (
    InventoryImportJobService,
    run_inventory_import_job,
)
from app.services.inventory.stock_service import StockService

router = APIRouter(
    prefix="/inventory",
    tags=["inventory"],
    dependencies=[Depends(require_role("admin", "stock"))],
)

REQUIRED_COLUMNS = {"sku", "name", "category", "price", "cost", "stock", "stock_min"}
SUPPORTED_FILE_TYPES = {"csv", "xlsx"}


def _detect_file_type(filename: str | None) -> str:
    if not filename:
        raise HTTPException(status_code=400, detail="Archivo invalido")
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o XLSX")


def _encode_cursor(item: StockMovementOut) -> str:
    payload = f"{item.created_at.isoformat()}|{item.id}"
    return base64.urlsafe_b64encode(payload.encode("utf-8")).decode("utf-8")


def _decode_cursor(cursor: str) -> tuple[datetime, int]:
    try:
        decoded = base64.urlsafe_b64decode(cursor.encode("utf-8")).decode("utf-8")
        timestamp, raw_id = decoded.rsplit("|", 1)
        return datetime.fromisoformat(timestamp), int(raw_id)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Cursor invalido") from exc


def _validate_date(value: str | None, field_name: str) -> str | None:
    if value is None:
        return None
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Filtro '{field_name}' invalido. Formato esperado YYYY-MM-DD",
        ) from exc
    return value


async def _read_upload_rows(file: UploadFile) -> list[dict]:
    filename = file.filename or ""
    file_type = _detect_file_type(filename)
    rows: list[dict] = []

    if file_type == "csv":
        content = await file.read()
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="CSV sin encabezados")
        header = {h.strip() for h in reader.fieldnames if h}
        if not REQUIRED_COLUMNS.issubset(header):
            missing = ", ".join(sorted(REQUIRED_COLUMNS - header))
            raise HTTPException(status_code=400, detail=f"Faltan columnas: {missing}")
        rows = [row for row in reader]
    elif file_type == "xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail="Instale openpyxl para XLSX"
            ) from exc
        data = await file.read()
        workbook = load_workbook(filename=BytesIO(data))
        worksheet = workbook.active
        if worksheet is None:
            raise HTTPException(status_code=400, detail="XLSX sin hoja activa")
        header_rows = list(worksheet.iter_rows(min_row=1, max_row=1))
        if not header_rows:
            raise HTTPException(status_code=400, detail="XLSX sin encabezados")
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in header_rows[0]
        ]
        if not any(headers):
            raise HTTPException(status_code=400, detail="XLSX sin encabezados")
        header = {h for h in headers if h}
        if not REQUIRED_COLUMNS.issubset(header):
            missing = ", ".join(sorted(REQUIRED_COLUMNS - header))
            raise HTTPException(status_code=400, detail=f"Faltan columnas: {missing}")
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({headers[index]: row[index] for index in range(len(headers))})
    return rows


async def _write_upload_to_temp_file(
    file: UploadFile, file_type: str
) -> tuple[str, int]:
    suffix = ".csv" if file_type == "csv" else ".xlsx"
    handler = tempfile.NamedTemporaryFile(
        prefix="inventory-import-", suffix=suffix, delete=False
    )
    max_bytes = max(1, settings.inventory_import_max_file_size_mb) * 1024 * 1024
    size = 0
    try:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > max_bytes:
                raise HTTPException(
                    status_code=413,
                    detail=f"Archivo supera el limite de {settings.inventory_import_max_file_size_mb}MB",
                )
            handler.write(chunk)
    except Exception:
        handler.close()
        try:
            os.remove(handler.name)
        except FileNotFoundError:
            pass
        raise
    finally:
        handler.close()
        await file.seek(0)
    return handler.name, size


def _assert_job_access(job: InventoryImportJob, current_user) -> None:
    if current_user.role == "admin":
        return
    if job.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Sin permisos")


@router.post(
    "/movement",
    response_model=StockMovementOut,
    status_code=201,
    dependencies=[Depends(require_permission("inventory.write"))],
)
async def create_movement(
    data: InventoryMovementCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    service = StockService(db, current_user)
    return await service.create_movement(data)


@router.get("/template", dependencies=[Depends(require_permission("inventory.read"))])
async def download_template():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["sku", "name", "category", "price", "cost", "stock", "stock_min"])
    writer.writerow(["BK-001", "Libro ejemplo", "Ficcion", "25.90", "12.50", "10", "2"])
    return PlainTextResponse(output.getvalue(), media_type="text/csv")


@router.get(
    "/template/xlsx", dependencies=[Depends(require_permission("inventory.read"))]
)
async def download_template_xlsx():
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail="Instale openpyxl para XLSX"
        ) from exc
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise HTTPException(
            status_code=500, detail="No se pudo crear la hoja de trabajo"
        )
    worksheet.append(["sku", "name", "category", "price", "cost", "stock", "stock_min"])
    worksheet.append(
        ["BK-001", "Libro ejemplo", "Ficcion", "25.90", "12.50", "10", "2"]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_template.xlsx"},
    )


@router.post("/upload", dependencies=[Depends(require_permission("inventory.write"))])
async def upload_inventory(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows = await _read_upload_rows(file)
    service = StockService(db, current_user)
    return await service.bulk_import(rows)


@router.post(
    "/import-jobs",
    response_model=InventoryImportJobOut,
    status_code=201,
    dependencies=[Depends(require_permission("inventory.write"))],
)
async def create_import_job(
    request: Request,
    background_tasks: BackgroundTasks,
    batch_size: int | None = Query(default=None, ge=1),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    file_type = _detect_file_type(file.filename)
    temp_path, _size = await _write_upload_to_temp_file(file, file_type)
    request_id = getattr(request.state, "request_id", None)
    service = InventoryImportJobService(db, current_user, request_id=request_id)
    job = await service.create_job(
        filename=file.filename or f"import.{file_type}",
        file_type=file_type,
        batch_size=batch_size or settings.inventory_import_default_batch_size,
    )
    background_tasks.add_task(run_inventory_import_job, job.id, temp_path)
    return job


@router.get(
    "/import-jobs/{job_id}",
    response_model=InventoryImportJobOut,
    dependencies=[Depends(require_permission("inventory.read"))],
)
async def get_import_job(
    job_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    service = InventoryImportJobService(db, current_user)
    job = await service.get_job(job_id)
    _assert_job_access(job, current_user)
    return job


@router.get(
    "/import-jobs/{job_id}/errors",
    dependencies=[Depends(require_permission("inventory.read"))],
)
async def get_import_job_errors(
    job_id: int,
    format: str = Query(default="json", pattern="^(json|csv)$"),
    limit: int = Query(default=500, ge=1, le=5000),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    service = InventoryImportJobService(db, current_user)
    job, items = await service.get_job_errors(job_id, limit=limit)
    _assert_job_access(job, current_user)
    if format == "csv":
        content = await service.export_job_errors_csv(job_id)
        return PlainTextResponse(
            content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=import_job_{job_id}_errors.csv"
            },
        )
    return InventoryImportJobErrorListOut(
        job_id=job.id,
        total_errors=job.error_rows,
        items=[InventoryImportJobErrorOut.model_validate(item) for item in items],
    )


@router.get(
    "/kardex/{product_id}",
    response_model=KardexPageOut,
    dependencies=[Depends(require_permission("inventory.read"))],
)
async def get_kardex(
    product_id: int,
    limit: int = Query(default=100, ge=1, le=500),
    cursor: str | None = Query(default=None),
    from_date: str | None = Query(default=None, alias="from"),
    to_date: str | None = Query(default=None, alias="to"),
    movement_type: str | None = Query(default=None, alias="type"),
    db: AsyncSession = Depends(get_db),
):
    safe_from = _validate_date(from_date, "from")
    safe_to = _validate_date(to_date, "to")

    stmt = select(StockMovement).where(StockMovement.product_id == product_id)
    if safe_from:
        stmt = stmt.where(func.date(StockMovement.created_at) >= safe_from)
    if safe_to:
        stmt = stmt.where(func.date(StockMovement.created_at) <= safe_to)
    if movement_type:
        normalized = movement_type.strip().upper()
        if normalized not in {"IN", "OUT", "ADJ", "TRF"}:
            raise HTTPException(status_code=400, detail="Filtro 'type' invalido")
        stmt = stmt.where(StockMovement.type == normalized)
    if cursor:
        cursor_created_at, cursor_id = _decode_cursor(cursor)
        stmt = stmt.where(
            or_(
                StockMovement.created_at < cursor_created_at,
                and_(
                    StockMovement.created_at == cursor_created_at,
                    StockMovement.id < cursor_id,
                ),
            )
        )

    result = await db.execute(
        stmt.order_by(StockMovement.created_at.desc(), StockMovement.id.desc()).limit(
            limit + 1
        )
    )
    rows = list(result.scalars().all())
    has_more = len(rows) > limit
    visible = rows[:limit]
    items = [StockMovementOut.model_validate(row) for row in visible]
    next_cursor = _encode_cursor(items[-1]) if has_more and items else None
    return KardexPageOut(
        items=items, limit=limit, has_more=has_more, next_cursor=next_cursor
    )
