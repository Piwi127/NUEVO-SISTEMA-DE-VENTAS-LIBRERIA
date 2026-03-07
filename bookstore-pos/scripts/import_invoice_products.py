from __future__ import annotations

import argparse
import asyncio
import csv
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.audit import log_event
from app.core.stock import apply_stock_delta, require_default_warehouse_id
from app.db.session import AsyncSessionLocal
from app.models.inventory import StockMovement
from app.models.product import Product


SOURCE_TAG = "fuente:productos_boletas_extraidos.xlsx"
EXPECTED_HEADERS = {
    "Documento": "document",
    "Item": "item",
    "Cantidad": "quantity",
    "UM": "unit_measure",
    "Material": "material",
    "Descripcion": "description",
    "Precio Unitario": "unit_price",
    "Total": "line_total",
}


@dataclass
class NormalizedProduct:
    source_document: str
    source_item: str
    unit_measure: str
    sku: str
    name: str
    category: str
    tags: str
    price: float
    cost: float
    stock: int
    stock_min: int


def normalize_text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else ("%s" % value).rstrip("0").rstrip(".")
    return normalize_text(value)


def parse_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return round(float(value), 4)


def parse_int(value: object) -> int:
    if value in (None, ""):
        return 0
    return int(round(float(value)))


def classify_category(name: str) -> str:
    text = name.upper()

    if any(token in text for token in ("PLUMON", "LAPIZ", "LAPICERO", "MARCADOR", "RESALTADOR", "BORRADOR", "TAJADOR", "LIMPIATIPOS", "CORRECTOR")):
        return "Escritura"
    if any(token in text for token in ("PEGAMENTO", "SILICONA", "ADHESIVO", "MASKING", "SCOTCH", "CINTA", "GOMA")):
        return "Manualidades"
    if any(token in text for token in ("PORTAPAPEL", "ARCHIVADOR", "CLIP", "GRAPA", "ENGRAPADOR", "SEPARADOR", "SOBRE", "FILE")):
        return "Oficina"
    if any(token in text for token in ("FORRO", "CUADERNO", "REGLA", "COMPAS", "TRANSPORTADOR", "MOCHILA", "LONCHERA")):
        return "Escolar"
    if any(token in text for token in ("ALCOHOL", "JABON", "LIMPIA", "DESINFECT", "PAÑO", "PANO")):
        return "Aseo"
    return "Papeleria"


def build_tags(document: str, unit_measure: str) -> str:
    parts = [SOURCE_TAG]
    if document:
        parts.append(f"documento:{document}")
    if unit_measure:
        parts.append(f"um:{unit_measure}")
    return ", ".join(parts)


def merge_tags(existing: str, extra: str) -> str:
    current = [part.strip() for part in str(existing or "").split(",") if part.strip()]
    for part in [piece.strip() for piece in extra.split(",") if piece.strip()]:
        if part not in current:
            current.append(part)
    return ", ".join(current)


def load_products_from_excel(path: Path) -> list[NormalizedProduct]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Productos" not in workbook.sheetnames:
        raise ValueError("No se encontro la hoja 'Productos'")
    sheet = workbook["Productos"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_text(value) for value in rows[0]]
    missing = [header for header in EXPECTED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Faltan columnas en Excel: {', '.join(missing)}")

    header_index = {EXPECTED_HEADERS[header]: headers.index(header) for header in EXPECTED_HEADERS}
    normalized: dict[str, NormalizedProduct] = {}

    for excel_row in rows[1:]:
        document = normalize_text(excel_row[header_index["document"]])
        item = normalize_text(excel_row[header_index["item"]])
        unit_measure = normalize_text(excel_row[header_index["unit_measure"]])
        sku = normalize_code(excel_row[header_index["material"]])
        name = normalize_text(excel_row[header_index["description"]])
        quantity = parse_int(excel_row[header_index["quantity"]])
        unit_price = parse_float(excel_row[header_index["unit_price"]])

        if not sku or not name:
            continue

        row = NormalizedProduct(
            source_document=document,
            source_item=item,
            unit_measure=unit_measure,
            sku=sku,
            name=name,
            category=classify_category(name),
            tags=build_tags(document, unit_measure),
            price=unit_price,
            cost=unit_price,
            stock=max(quantity, 0),
            stock_min=0,
        )

        existing = normalized.get(sku)
        if existing is None:
            normalized[sku] = row
            continue

        existing.stock += row.stock
        existing.cost = row.cost
        existing.price = row.price
        existing.tags = merge_tags(existing.tags, row.tags)
        if row.source_document and row.source_document not in existing.source_document:
            existing.source_document = f"{existing.source_document} | {row.source_document}".strip(" |")

    return list(normalized.values())


def write_normalized_csv(path: Path, rows: list[NormalizedProduct]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "sku",
            "name",
            "category",
            "tags",
            "price",
            "cost",
            "stock",
            "stock_min",
            "source_document",
            "unit_measure",
        ])
        for row in rows:
            writer.writerow([
                row.sku,
                row.name,
                row.category,
                row.tags,
                row.price,
                row.cost,
                row.stock,
                row.stock_min,
                row.source_document,
                row.unit_measure,
            ])


async def import_products(rows: list[NormalizedProduct], dry_run: bool) -> dict[str, object]:
    created = 0
    updated = 0
    stock_units = 0
    overlaps: list[str] = []

    async with AsyncSessionLocal() as db:
        warehouse_id = await require_default_warehouse_id(db)

        for row in rows:
            result = await db.execute(select(Product).where(Product.sku == row.sku))
            product = result.scalar_one_or_none()

            if product is None:
                product = Product(
                    sku=row.sku,
                    name=row.name,
                    category=row.category,
                    tags=row.tags,
                    price=row.price,
                    sale_price=row.price,
                    cost=row.cost,
                    unit_cost=row.cost,
                    cost_total=row.cost,
                    cost_qty=1,
                    direct_costs_breakdown="{}",
                    direct_costs_total=0,
                    desired_margin=0,
                    stock=0,
                    stock_min=row.stock_min,
                )
                db.add(product)
                await db.flush()
                created += 1
            else:
                overlaps.append(row.sku)
                updated += 1
                product.cost = row.cost
                product.unit_cost = row.cost
                product.cost_total = row.cost
                product.cost_qty = 1
                product.direct_costs_breakdown = "{}"
                product.direct_costs_total = 0
                if not normalize_text(product.category):
                    product.category = row.category
                if not normalize_text(product.tags):
                    product.tags = row.tags
                else:
                    product.tags = merge_tags(product.tags, row.tags)
                if not normalize_text(product.name):
                    product.name = row.name
                if float(product.price or 0) <= 0:
                    product.price = row.price
                if float(product.sale_price or 0) <= 0:
                    product.sale_price = row.price

            if row.stock > 0:
                stock_units += row.stock
                if not dry_run:
                    await apply_stock_delta(db, product.id, row.stock, warehouse_id)
                    db.add(
                        StockMovement(
                            product_id=product.id,
                            type="IN",
                            qty=row.stock,
                            ref=f"XLSX:{row.source_document or 'SIN-DOC'}:{row.source_item or row.sku}"[:100],
                        )
                    )

        if dry_run:
            await db.rollback()
        else:
            await db.commit()

    return {
        "rows": len(rows),
        "created": created,
        "updated": updated,
        "stock_units": stock_units,
        "overlaps": overlaps,
    }


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Importa productos desde un Excel de boletas al esquema del proyecto.")
    parser.add_argument("excel_path", type=Path, help="Ruta al archivo XLSX de origen")
    parser.add_argument("--dry-run", action="store_true", help="Analiza y simula sin grabar en la base")
    parser.add_argument("--csv-out", type=Path, default=None, help="Ruta opcional para guardar el CSV normalizado")
    return parser


async def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()

    rows = load_products_from_excel(args.excel_path)
    if args.csv_out:
        write_normalized_csv(args.csv_out, rows)

    summary = await import_products(rows, dry_run=args.dry_run)
    print(f"ROWS={summary['rows']}")
    print(f"CREATED={summary['created']}")
    print(f"UPDATED={summary['updated']}")
    print(f"STOCK_UNITS={summary['stock_units']}")
    print(f"OVERLAPS={','.join(summary['overlaps']) if summary['overlaps'] else '-'}")
    if args.csv_out:
        print(f"CSV_OUT={args.csv_out}")
    print(f"DRY_RUN={args.dry_run}")


if __name__ == "__main__":
    asyncio.run(main())


